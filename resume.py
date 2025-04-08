import os
import re
import json
import time
import random
import zipfile
import argparse
import requests
from openpyxl import load_workbook
from seleniumwire import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import StaleElementReferenceException


def scroll_to_center(driver, element):
    driver.execute_script(
        "arguments[0].scrollIntoView({block: 'center', inline: 'center'});", element
    )


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", help="Путь к Excel шаблону", required=True)
    parser.add_argument(
        "--proxy", help="Прокси в формате host:port:user:pass", required=True
    )
    parser.add_argument("--cookies", help="Путь к файлу с куки", required=True)
    return parser.parse_args()


class DataExtractor:
    def __init__(self, template_path, sheet_name):
        self.template_path = template_path
        self.sheet_name = sheet_name

    def extract_data(self):
        wb = load_workbook(self.template_path, data_only=True)
        ws = wb[self.sheet_name]
        data = {
            "profession": ws["B1"].value,
            "education": ws["B3"].value,
            "institution": ws["B4"].value,
            "faculty_specialty": ws["B5"].value,
            "graduation_year": ws["B6"].value,
            "job_description": ws["A43"].value,
            "skills": [],
            "companies": [],
        }
        for row in ws.iter_rows(min_row=13, max_row=27, min_col=1, max_col=2):
            skill = row[0].value
            level = row[1].value
            if skill:
                data["skills"].append({"skill": skill, "level": level})
        for col in range(2, 6):
            name = ws.cell(row=31, column=col).value
            if not name:
                continue
            company = {
                "name": name,
                "position": ws.cell(row=32, column=col).value,
                "description": ws.cell(row=33, column=col).value,
                "start_month": ws.cell(row=34, column=col).value,
                "start_year": ws.cell(row=35, column=col).value,
                "end_month": ws.cell(row=36, column=col).value,
                "end_year": ws.cell(row=37, column=col).value,
            }
            if str(company["end_month"]).strip().upper() == "НВ":
                company["end_month"] = "НАСТОЯЩЕЕ ВРЕМЯ"
                company["end_year"] = None
            data["companies"].append(company)
        return data



class ResumeFiller:
    def __init__(self, driver):
        self.driver = driver
        self.wait = WebDriverWait(self.driver, 10)

    def fill_company_data(self, company, index=0):
        if not (
            company.get("end_year") is None
            and company.get("end_month", "").strip().upper() == "НАСТОЯЩЕЕ ВРЕМЯ"
        ):
            try:
                checkbox = self.wait.until(
                    EC.element_to_be_clickable((
                        By.CSS_SELECTOR,
                        "label[data-qa='cell']",
                    ))
                )
                scroll_to_center(self.driver, checkbox)
                self.driver.execute_script("arguments[0].click();", checkbox)
                time.sleep(0.2)
            except Exception as e:
                print(f"[{index}] Ошибка при клике по галочке 'Сейчас работаю':", e)
        month_mapping = {
            "Январь": "01",
            "Февраль": "02",
            "Март": "03",
            "Апрель": "04",
            "Май": "05",
            "Июнь": "06",
            "Июль": "07",
            "Август": "08",
            "Сентябрь": "09",
            "Октябрь": "10",
            "Ноябрь": "11",
            "Декабрь": "12",
        }
        company_selector = (
            f"input[data-qa='resume-profile-experience-specific-company-input-{index}']"
        )
        position_selector = f"input[data-qa='resume-profile-experience-specific-position-input-{index}']"
        responsibilities_selector = f"textarea[data-qa='resume-profile-experience-specific-responsibilities-input-{index}']"
        datestart_year_selector = f"input[data-qa='resume-profile-experience-specific-datestart-year-input-{index}']"
        dateend_year_selector = f"input[data-qa='resume-profile-experience-specific-dateend-year-input-{index}']"
        try:
            company_input = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, company_selector))
            )
            scroll_to_center(self.driver, company_input)
            company_input.clear()
            time.sleep(0.2)
            company_names_raw = company.get("name", "")
            company_names = [
                name.strip() for name in company_names_raw.split(",") if name.strip()
            ]
            selected_name = random.choice(company_names) if company_names else ""
            company_input.send_keys(selected_name)
        except Exception as e:
            print(f"[{index}] Ошибка заполнения названия компании:", e)
        try:
            position_input = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, position_selector))
            )
            scroll_to_center(self.driver, position_input)
            position_input.clear()
            time.sleep(0.2)
            position_input.send_keys(company.get("position", ""))
        except Exception as e:
            print(f"[{index}] Ошибка заполнения должности:", e)
        try:
            experience_title = self.wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "h1[data-qa='title']"))
            )
            scroll_to_center(self.driver, experience_title)
            try:
                experience_title.click()
                time.sleep(0.2)
            except Exception:
                try:
                    h2_title = self.wait.until(
                        EC.element_to_be_clickable((
                            By.CSS_SELECTOR,
                            "h2[data-qa='title']",
                        ))
                    )
                    scroll_to_center(self.driver, h2_title)
                    h2_title.click()
                    time.sleep(0.2)
                except Exception as e:
                    print(f"[{index}] Не удалось кликнуть по h2:", e)
                    return
        except Exception as e:
            print(f"[{index}] Заголовок h1 не найден:", e)
            return
        try:
            responsibilities_input = self.wait.until(
                EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    responsibilities_selector,
                ))
            )
            scroll_to_center(self.driver, responsibilities_input)
            responsibilities_input.clear()
            time.sleep(0.2)
            responsibilities_input.send_keys(company.get("description", ""))
        except Exception as e:
            print(f"[{index}] Ошибка заполнения описания:", e)
        start_year = str(company.get("start_year", "")).strip()
        try:
            start_year_input = self.wait.until(
                EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    datestart_year_selector,
                ))
            )
            scroll_to_center(self.driver, start_year_input)
            start_year_input.clear()
            time.sleep(0.2)
            start_year_input.send_keys(start_year)
        except Exception as e:
            print(f"[{index}] Ошибка заполнения года начала:", e)
        time.sleep(0.2)
        start_month = company.get("start_month", "").strip()
        try:
            activators = self.wait.until(
                EC.presence_of_all_elements_located((
                    By.CSS_SELECTOR,
                    "div[data-qa='magritte-select-activator']",
                ))
            )
            if len(activators) >= 1:
                scroll_to_center(self.driver, activators[0])
                activators[0].click()
                time.sleep(0.2)
                self.wait.until(
                    EC.visibility_of_element_located((
                        By.CSS_SELECTOR,
                        "div[data-qa='magritte-select-option-list']",
                    ))
                )
                start_month_number = month_mapping.get(start_month)
                if start_month_number:
                    option = self.wait.until(
                        EC.element_to_be_clickable((
                            By.CSS_SELECTOR,
                            f"div[data-qa='magritte-select-option-list'] label[data-magritte-select-option='{start_month_number}']",
                        ))
                    )
                    scroll_to_center(self.driver, option)
                    option.click()
                    time.sleep(0.2)
                else:
                    print(
                        f"[{index}] Не найден номер для месяца начала: '{start_month}'"
                    )
            else:
                print(f"[{index}] Активатор для месяца начала не найден.")
        except Exception as e:
            print(f"[{index}] Ошибка при выборе месяца начала:", e)
        if (
            company.get("end_year") is None
            and company.get("end_month", "").strip().upper() == "НАСТОЯЩЕЕ ВРЕМЯ"
        ):
            pass
        else:
            end_year = (
                str(company.get("end_year", "")).strip()
                if company.get("end_year") is not None
                else ""
            )
            try:
                end_year_input = self.wait.until(
                    EC.presence_of_element_located((
                        By.CSS_SELECTOR,
                        dateend_year_selector,
                    ))
                )
                scroll_to_center(self.driver, end_year_input)
                end_year_input.clear()
                time.sleep(0.2)
                end_year_input.send_keys(end_year)
            except Exception as e:
                print(f"[{index}] Ошибка заполнения года окончания:", e)
            try:
                end_month = company.get("end_month", "").strip()
                activators = self.wait.until(
                    EC.presence_of_all_elements_located((
                        By.CSS_SELECTOR,
                        "div[data-qa='magritte-select-activator']",
                    ))
                )
                if len(activators) >= 1:
                    scroll_to_center(self.driver, activators[1])
                    activators[1].click()
                    time.sleep(0.2)
                    self.wait.until(
                        EC.visibility_of_element_located((
                            By.CSS_SELECTOR,
                            "div[data-qa='magritte-select-option-list']",
                        ))
                    )
                    end_month_number = month_mapping.get(end_month)
                    if end_month_number:
                        option = self.wait.until(
                            EC.element_to_be_clickable((
                                By.CSS_SELECTOR,
                                f"div[data-qa='magritte-select-option-list'] label[data-magritte-select-option='{end_month_number}']",
                            ))
                        )
                        scroll_to_center(self.driver, option)
                        option.click()
                        time.sleep(0.2)
                    else:
                        print(
                            f"[{index}] Не найден номер для месяца окончания: '{end_month}'"
                        )
                else:
                    print(f"[{index}] Активатор для месяца окончания не найден.")
            except Exception as e:
                print(f"[{index}] Ошибка при выборе месяца окончания:", e)

    def fill_companies(self, companies):
        for index, company in enumerate(companies):
            if index > 0:
                try:
                    add_button = self.wait.until(
                        EC.element_to_be_clickable((
                            By.CSS_SELECTOR,
                            "button[data-qa='list-add']",
                        ))
                    )
                    scroll_to_center(self.driver, add_button)
                    try:
                        add_button.click()
                        time.sleep(0.2)
                    except Exception as e:
                        if "Other element would receive the click" in str(e):
                            try:
                                close_modal_btn = self.wait.until(
                                    EC.element_to_be_clickable((
                                        By.CSS_SELECTOR,
                                        "[data-qa='actions-container-modal actions-container-modal-or-vertical'] button",
                                    ))
                                )
                                scroll_to_center(self.driver, close_modal_btn)
                                close_modal_btn.click()
                                time.sleep(0.2)
                                self.wait.until(
                                    EC.invisibility_of_element_located((
                                        By.CSS_SELECTOR,
                                        "div[data-qa='modal-overlay']",
                                    ))
                                )
                                add_button = self.wait.until(
                                    EC.element_to_be_clickable((
                                        By.CSS_SELECTOR,
                                        "button[data-qa='list-add']",
                                    ))
                                )
                                scroll_to_center(self.driver, add_button)
                                add_button.click()
                                time.sleep(0.2)
                            except Exception as inner_e:
                                print(
                                    f"[{index}] Ошибка при закрытии модалки:", inner_e
                                )
                                return
                        else:
                            print(
                                f"[{index}] Ошибка при клике по кнопке добавления:", e
                            )
                            return
                except Exception as e:
                    print(f"[{index}] Ошибка при добавлении компании:", e)
                    return
            self.fill_company_data(company, index=index)
        try:
            primary_actions_button = self.wait.until(
                EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    "div[data-qa='primary-actions'] button",
                ))
            )
            scroll_to_center(self.driver, primary_actions_button)
            primary_actions_button.click()
            time.sleep(0.2)
        except Exception as e:
            print("Ошибка при нажатии кнопки 'primary-actions':", e)
            return

    def fill_resume(self, data):
        self.driver.get("https://hh.ru/profile/resume/professional_role")
        try:
            job_select = self.wait.until(
                EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    "div[data-qa='resume-profile-card-select-job']",
                ))
            )
            scroll_to_center(self.driver, job_select)
            job_select.click()
            time.sleep(0.2)
        except Exception as e:
            print("Ошибка при выборе профессии:", e)
            return
        if not wait_for_h1_title(self.driver, "Выберите или укажите профессию"):
            return
        profession = data.get("profession", "")
        try:
            position_input = self.wait.until(
                EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    "input[data-qa='resume-profile-position-input']",
                ))
            )
            scroll_to_center(self.driver, position_input)
            position_input.clear()
            position_input.send_keys(profession)
            position_input.click()
            time.sleep(0.2)
        except Exception as e:
            print("Ошибка при заполнении поля профессии:", e)
            return
        try:
            self.wait.until(
                EC.visibility_of_element_located((
                    By.CSS_SELECTOR,
                    "div[data-qa='suggest-drop']",
                ))
            )
            suggest_items = self.wait.until(
                EC.visibility_of_all_elements_located((
                    By.CSS_SELECTOR,
                    "div[data-qa='suggest-item-cell']",
                ))
            )
            found = False
            for item in suggest_items:
                if item.text.strip().lower() == profession.strip().lower():
                    scroll_to_center(self.driver, item)
                    item.click()
                    time.sleep(0.2)
                    found = True
                    break
            if not found:
                print("Подсказка по профессии не найдена.")
                return
        except Exception as e:
            print("Ошибка при выборе подсказки по профессии:", e)
            return
        try:
            final_button = self.wait.until(
                EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    ".ResumeProfileFront-ReactRoot [class*='___footer'] [class*='___footer-right'] button",
                ))
            )
            scroll_to_center(self.driver, final_button)
            final_button.click()
            time.sleep(0.2)
        except Exception as e:
            print("Ошибка при первом клике по финальной кнопке:", e)
            return
        if not wait_for_h1_title(self.driver, "Заполните основную информацию"):
            return
        try:
            final_button = self.wait.until(
                EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    ".ResumeProfileFront-ReactRoot [class*='___footer'] [class*='___footer-right'] button",
                ))
            )
            scroll_to_center(self.driver, final_button)
            final_button.click()
            time.sleep(0.2)
        except Exception as e:
            print("Ошибка при втором клике по финальной кнопке:", e)
            return
        if not wait_for_h1_title(self.driver, "Какое у вас образование?"):
            return
        try:
            education_value = data.get("education", "")
            label_element = self.wait.until(
                EC.presence_of_element_located((
                    By.XPATH,
                    f"//label[input[contains(@aria-label, '{education_value}')]]",
                ))
            )
            scroll_to_center(self.driver, label_element)
            label_element.click()
            time.sleep(0.2)
        except Exception as e:
            print("Ошибка при выборе уровня образования:", e)
            return
        if not wait_for_h1_title(self.driver, "Какое учебное заведение окончили?"):
            return
        try:
            univ_input = self.wait.until(
                EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    "input[data-qa='resume-profile-education-specific-university-input-0']",
                ))
            )
            scroll_to_center(self.driver, univ_input)
            univ_input.clear()
            univ_input.send_keys(data.get("institution", ""))
            faculty_input = self.wait.until(
                EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    "input[data-qa='resume-profile-education-specific-faculty-input-0']",
                ))
            )
            scroll_to_center(self.driver, faculty_input)
            faculty_input.clear()
            faculty_input.send_keys(data.get("faculty_specialty", ""))
            spec_input = self.wait.until(
                EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    "input[data-qa='resume-profile-education-specific-spec-input-0']",
                ))
            )
            scroll_to_center(self.driver, spec_input)
            spec_input.clear()
            spec_input.send_keys(data.get("faculty_specialty", ""))
            year_input = self.wait.until(
                EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    "input[data-qa='resume-profile-primaryEducation-specific-year-input-0']",
                ))
            )
            scroll_to_center(self.driver, year_input)
            year_input.clear()
            year_input.send_keys(str(data.get("graduation_year", "")))
        except Exception as e:
            print("Ошибка при заполнении полей учебного заведения:", e)
            return
        try:
            final_button = self.wait.until(
                EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    ".ResumeProfileFront-ReactRoot [class*='___footer'] [class*='___footer-right'] button",
                ))
            )
            scroll_to_center(self.driver, final_button)
            final_button.click()
            time.sleep(0.2)
        except Exception as e:
            print("Ошибка при третьем клике по финальной кнопке:", e)
            return
        if not wait_for_h1_title(self.driver, "Какими навыками обладаете?"):
            return
        try:
            chips_input = self.wait.until(
                EC.presence_of_element_located((
                    By.CSS_SELECTOR,
                    "input[data-qa='chips-trigger-input']",
                ))
            )
            scroll_to_center(self.driver, chips_input)
            for skill_item in data.get("skills", []):
                skill_text = skill_item.get("skill", "")
                if skill_text:
                    chips_input.send_keys(skill_text)
                    chips_input.send_keys(Keys.ENTER)
                    time.sleep(0.2)
        except Exception as e:
            print("Ошибка при добавлении навыков:", e)
            return
        try:
            final_button = self.wait.until(
                EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    ".ResumeProfileFront-ReactRoot [class*='___footer'] [class*='___footer-right'] button",
                ))
            )
            scroll_to_center(self.driver, final_button)
            final_button.click()
            time.sleep(0.2)
        except Exception as e:
            print("Ошибка при четвертом клике по финальной кнопке:", e)
            return
        if not wait_for_h1_title(self.driver, "На каких уровнях владеете навыками"):
            return

        def normalize(text):
            return " ".join(text.split()).lower()

        try:
            skill_containers = self.driver.find_elements(
                By.CSS_SELECTOR, "div[data-qa='skill']"
            )
            for container in skill_containers:
                try:
                    skill_name_elem = container.find_element(
                        By.CSS_SELECTOR, "div[data-qa='skillName']"
                    )
                    actual_skill = normalize(skill_name_elem.text)
                except Exception:
                    continue
                for json_skill in data.get("skills", []):
                    expected_skill = normalize(json_skill.get("skill", ""))
                    expected_level = str(json_skill.get("level")).strip()
                    if actual_skill == expected_skill:
                        labels = container.find_elements(By.CSS_SELECTOR, "label")
                        clicked = False
                        for label in labels:
                            try:
                                child_div = label.find_element(
                                    By.CSS_SELECTOR, "div[data-qa^='skill-level-']"
                                )
                                level_attr = child_div.get_attribute("data-qa")
                                match = re.search(r"skill-level-(\d+)", level_attr)
                                if match and match.group(1) == expected_level:
                                    scroll_to_center(self.driver, label)
                                    label.click()
                                    time.sleep(0.2)
                                    clicked = True
                                    break
                            except Exception:
                                continue
                        if not clicked:
                            print(
                                f"Label с уровнем {expected_level} не найден для навыка '{expected_skill}'"
                            )
                        break
        except Exception as e:
            print("Ошибка при поиске навыков:", e)
        try:
            final_button = self.wait.until(
                EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    ".ResumeProfileFront-ReactRoot [class*='___footer'] [class*='___footer-right'] button",
                ))
            )
            scroll_to_center(self.driver, final_button)
            final_button.click()
            time.sleep(0.2)
        except Exception as e:
            print("Ошибка при пятом клике по финальной кнопке:", e)
            return
        companies = data.get("companies", [])
        if companies:
            self.fill_companies(companies)
        else:
            print("Компании не обнаружены в данных.")
        try:
            final_button = self.wait.until(
                EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    ".ResumeProfileFront-ReactRoot [class*='___footer'] [class*='___footer-right'] button:not([data-qa='list-add'])",
                ))
            )
            scroll_to_center(self.driver, final_button)
            final_button.click()
            time.sleep(0.2)
        except Exception as e:
            print("Ошибка при первом клике по финальной кнопке:", e)
            return


def wait_for_h1_title(driver, expected_text, retries=3, delay=1):
    for attempt in range(retries):
        try:
            WebDriverWait(driver, 30).until(
                EC.text_to_be_present_in_element(
                    (By.CSS_SELECTOR, "h1[data-qa='title']"), expected_text
                )
            )
            return True
        except Exception:
            time.sleep(delay)
    print(f"Заголовок '{expected_text}' не найден после {retries} попыток.")
    return False


def create_driver_with_proxy(proxy_arg):
    proxy_host, proxy_port, proxy_user, proxy_pass = proxy_arg.split(":")
    seleniumwire_options = {
        'proxy': {
            'http': f'http://{proxy_user}:{proxy_pass}@{proxy_host}:{proxy_port}',
            'https': f'http://{proxy_user}:{proxy_pass}@{proxy_host}:{proxy_port}',
            'no_proxy': 'localhost,127.0.0.1'
        },
        'verify_ssl': False,         # Отключаем проверку SSL сертификатов
        'disable_capture': True,     # Отключаем перехват HTTPS-трафика (и генерацию своего сертификата)
        'connection_interceptor': False,
        'enable_har': False,
        'disable_encoding': True,
    }
    chrome_options = Options()
    chrome_options.add_argument("--ignore-certificate-errors")  # на всякий случай
    driver = webdriver.Chrome(seleniumwire_options=seleniumwire_options, options=chrome_options)
    return driver


def load_cookies(driver, cookies_file):
    driver.get("https://hh.ru")
    # Клик по кнопке принятия куки, если она появляется
    try:
        cookies_btn = WebDriverWait(driver, 5).until(
            EC.element_to_be_clickable((
                By.CSS_SELECTOR,
                '[data-qa="cookies-policy-informer-accept"]',
            ))
        )
        scroll_to_center(driver, cookies_btn)
        cookies_btn.click()
        time.sleep(0.5)
    except Exception as e:
        print("Кнопка принятия куки не найдена или не кликабельна:", e)
    time.sleep(1)
    with open(cookies_file, "r", encoding="utf-8") as f:
        cookies = json.load(f)
    for cookie in cookies:
        if "expiry" in cookie:
            cookie["expiry"] = int(cookie["expiry"])
        driver.add_cookie(cookie)
    driver.refresh()
    time.sleep(1)


# Дополните файл следующим образом:


def main(proxy_arg=None, template_arg=None, cookies_arg=None):
    if proxy_arg is None or template_arg is None or cookies_arg is None:
        args = parse_args()
        proxy_arg = args.proxy
        template_arg = args.template
        cookies_arg = args.cookies

    if not os.path.isfile(template_arg):
        raise Exception(f"Файл шаблона не найден: {template_arg}")

    sheet_name = "Лист1"
    extractor = DataExtractor(template_arg, sheet_name)
    data = extractor.extract_data()

    driver = create_driver_with_proxy(proxy_arg)
    load_cookies(driver, cookies_arg)

    filler = ResumeFiller(driver)
    filler.fill_resume(data)

    driver.quit()


if __name__ == "__main__":
    main()
