import argparse
import zipfile
import json
import random
import openpyxl
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("--cookies", help="Путь к файлу с куки", required=True)
    parser.add_argument("--template", help="Путь к файлу xlsx с данными (B1 - профессия, A43 - описание)", required=True)
    return parser.parse_args()

def scroll_to_center(driver, element):
    driver.execute_script("arguments[0].scrollIntoView({block: 'center', inline: 'center'});", element)

def clear_and_send(element, text):
    try:
        element.clear()
    except Exception:
        pass
    element.send_keys(Keys.CONTROL + "a")
    element.send_keys(Keys.DELETE)
    element.send_keys(text)


def create_driver_with_local_proxy():
    from selenium.webdriver.chrome.options import Options

    chrome_options = Options()
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--proxy-server=http://127.0.0.1:8899")  # локальный прокси

    driver = webdriver.Chrome(options=chrome_options)
    return driver

def load_cookies(driver, cookies_file):
    driver.get("https://hh.ru/profile/me")
    with open(cookies_file, "r", encoding="utf-8") as f:
        cookies = json.load(f)
    for cookie in cookies:
        if "expiry" in cookie:
            cookie["expiry"] = int(cookie["expiry"])
        driver.add_cookie(cookie)
    driver.refresh()

def read_template_data(template_path):
    wb = openpyxl.load_workbook(template_path, data_only=True)
    ws = wb["Лист1"]
    profesia = ws["B1"].value
    descr = ws["A43"].value
    return profesia, descr

def main(cookies_arg=None, template_arg=None):
    if cookies_arg is None or template_arg is None:
        args = parse_args()
        cookies_arg = cookies_arg or args.cookies
        template_arg = template_arg or args.template

    profesia, descr_text = read_template_data(template_arg)
    driver = create_driver_with_local_proxy()
    load_cookies(driver, cookies_arg)

    try:
        pref_area = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="profile-additional-card-preferred-work-area"]'))
        )
        scroll_to_center(driver, pref_area)
        pref_area.click()
    except Exception as e:
        print("Ошибка при клике по preferred work area:", e)

    try:
        switch_elem = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[role="switch"]'))
        )
        scroll_to_center(driver, switch_elem)
        switch_elem.click()
    except Exception as e:
        print("Ошибка при клике по switch:", e)

    try:
        modal_btn = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="modal-header-image"] button'))
        )
        scroll_to_center(driver, modal_btn)
        modal_btn.click()
    except Exception as e:
        print("Ошибка при клике по кнопке в модальном окне:", e)

    try:
        save_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, '//div[@data-qa="magritte-alert"]/..//button[contains(., "Сохранить")]'))
        )
        scroll_to_center(driver, save_button)
        save_button.click()
    except Exception as e:
        print("Ошибка при клике по кнопке 'Сохранить2':", e)

    # import time
    # time.sleep(100000)

    driver.get("https://hh.ru/applicant/resumes")
    try:
        status_link = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="job-search-status-change-link"]'))
        )
        scroll_to_center(driver, status_link)
        status_link.click()
    except Exception as e:
        print("Ошибка при клике по job-search-status-change-link:", e)

    try:
        active_search_btn = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="job-search-change_active_search"]'))
        )
        scroll_to_center(driver, active_search_btn)
        active_search_btn.click()
    except Exception as e:
        print("Ошибка при клике по job-search-change_active_search:", e)

    try:
        profession_elem = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, f'//h3[text()="{profesia}"]'))
        )
        scroll_to_center(driver, profession_elem)
        driver.execute_script("arguments[0].click();", profession_elem)
    except Exception as e:
        print("Ошибка при клике по h3 с текстом профессии:", e)

    try:
        about_me_elem = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="resume-progress-item-experience-skills"] a'))
        )
        scroll_to_center(driver, about_me_elem)
        about_me_elem.click()
    except Exception as e:
        try:
            skills_edit = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="resume-block-skills-edit"]'))
            )
            scroll_to_center(driver, skills_edit)
            skills_edit.click()
        except Exception as ex:
            try:
                skills_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[data-qa="suitable-vacancies-suggest-item-skills"]'))
                )
                scroll_to_center(driver, skills_button)
                skills_button.click()
            except Exception as exc:
                print("Ошибка при клике по [data-qa='suitable-vacancies-suggest-item-skills']", exc)

    try:
        textarea = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.TAG_NAME, "textarea"))
        )
        clear_and_send(textarea, descr_text)
    except Exception as e:
        print("Ошибка при очистке и заполнении textarea:", e)

    try:
        submit_btn = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="resume-submit"]'))
        )
        scroll_to_center(driver, submit_btn)
        submit_btn.click()
    except Exception as e:
        print("Ошибка при клике по кнопке 'resume-submit':", e)

    try:
        pos_edit = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="resume-block-position-edit"]'))
        )
        scroll_to_center(driver, pos_edit)
        pos_edit.click()
    except Exception as e:
        print("Ошибка при клике по resume-block-position-edit:", e)
    
    try:
        salary_input = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[data-qa="resume-salary-amount"]'))
        )
        scroll_to_center(driver, salary_input)
        salary = random.choice(range(60000, 90001, 5000))
        clear_and_send(salary_input, str(salary))
    except Exception as e:
        print("Ошибка при вводе зарплаты:", e)

    checkbox_selectors = [
        '[data-qa="createresume__employment-part"]',
        '[data-qa="createresume__employment-project"]',
        '[data-qa="createresume__employment-volunteer"]',
        '[data-qa="createresume__employment-probation"]',
        '[data-qa="createresume__workschedule-shift"]',
        '[data-qa="createresume__workschedule-flexible"]',
        '[data-qa="createresume__workschedule-remote"]',
        '[data-qa="createresume__workschedule-fly_in_fly_out"]',
    ]
    for selector in checkbox_selectors:
        try:
            checkbox = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, selector))
            )
            scroll_to_center(driver, checkbox)
            driver.execute_script("arguments[0].click();", checkbox)
        except Exception as e:
            print(f"Ошибка при клике по чекбоксу {selector}:", e)

    try:
        radio = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, 'input[type="radio"][value="ready"]'))
        )
        scroll_to_center(driver, radio)
        driver.execute_script("arguments[0].click();", radio)
    except Exception as e:
        print("Ошибка при клике по радио кнопке с value='ready':", e)

    try:
        submit_btn = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="resume-submit"]'))
        )
        scroll_to_center(driver, submit_btn)
        submit_btn.click()
    except Exception as e:
        print("Ошибка при клике по resume-submit:", e)

    try:
        contacts_edit = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="resume-block-contacts-edit"]'))
        )
        scroll_to_center(driver, contacts_edit)
        contacts_edit.click()
    except Exception as e:
        print("Ошибка при клике по resume-block-contacts-edit:", e)

    try:
        phone_comment0 = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[name="phone[0].comment"]'))
        )
        scroll_to_center(driver, phone_comment0)
        clear_and_send(phone_comment0, "ОЗНАКОМЬТЕСЬ С МОИМ СОПРОВОД. ПИСЬМОМ")
    except Exception as e:
        print("Ошибка при заполнении phone[0].comment:", e)

    try:
        phone_set_work = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="resume-phone-set-work"]'))
        )
        scroll_to_center(driver, phone_set_work)
        phone_set_work.click()
    except Exception as e:
        print("Ошибка при клике по resume-phone-set-work:", e)

    try:
        phone_formatted = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[name="phone[1].formatted"]'))
        )
        scroll_to_center(driver, phone_formatted)
        clear_and_send(phone_formatted, "+7 (933) 184-44-67")
    except Exception as e:
        print("Ошибка при заполнении phone[1].formatted:", e)

    try:
        phone_comment1 = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, 'input[name="phone[1].comment"]'))
        )
        scroll_to_center(driver, phone_comment1)
        clear_and_send(phone_comment1, "ЗВОНИТЬ СЮДА, ЗДЕСЬ ЖЕ ВОТСАП")
    except Exception as e:
        print("Ошибка при заполнении phone[1].comment:", e)

    try:
        preferred_contact = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="resume-preferred-contact-work_phone"]'))
        )
        scroll_to_center(driver, preferred_contact)
        preferred_contact.click()
    except Exception as e:
        print("Ошибка при клике по resume-preferred-contact-work_phone:", e)

    try:
        submit_btn = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.CSS_SELECTOR, '[data-qa="resume-submit"]'))
        )
        scroll_to_center(driver, submit_btn)
        submit_btn.click()
    except Exception as e:
        print("Ошибка при клике по кнопке Сохранить:", e)

    input("Нажмите Enter для завершения...")
    driver.quit()

if __name__ == "__main__":
    main()

